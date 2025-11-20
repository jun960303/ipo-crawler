# gui/app.py
import os
import csv
import threading
import ctypes
import tkinter as tk
from datetime import datetime

from crawler.base import init_db, get_upcoming_by_broker, get_connection
from crawler.ipo38 import crawl_38_all


class IPOApp:
    # 바탕화면 저장 체크
    def get_real_desktop_path():
        CSIDL_DESKTOP = 0x0000
        SHGFP_TYPE_CURRENT = 0
        buf = ctypes.create_unicode_buffer(260)
        ctypes.windll.shell32.SHGetFolderPathW(
            None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf
        )
        return buf.value

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("엄마 공모주 일정 수집기 v0.5")
        self.root.geometry("900x650")

        # 상태 플래그
        self.stop_flag = False
        self.spinner_running = False

        # DB 초기화
        init_db()

        self._build_ui()

    # ----------------------- UI 구성 -----------------------

    def _build_ui(self):
        # 상단 제목
        title = tk.Label(
            self.root,
            text="공모주 크롤링 프로그램",
            font=("맑은 고딕", 24, "bold"),
        )
        title.pack(pady=10)

        # 상태/로딩 라벨
        self.loading_label = tk.Label(self.root, text="", font=("맑은 고딕", 11))
        self.loading_label.pack(pady=5)

        # 버튼 영역
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        # 데이터 수집 버튼
        self.btn_collect = tk.Button(
            btn_frame,
            text="데이터 수집하기",
            width=20,
            command=self.run_collect_in_thread,
        )
        self.btn_collect.grid(row=0, column=0, padx=10, pady=5)

        # 중지 버튼
        self.btn_stop = tk.Button(
            btn_frame,
            text="크롤링 중지",
            width=20,
            command=self.stop_crawling,
        )
        self.btn_stop.grid(row=1, column=0, padx=10, pady=5)

        # 엑셀 내보내기
        btn_export = tk.Button(
            btn_frame,
            text="엑셀로 내보내기",
            width=20,
            command=self.export_to_excel,
        )
        btn_export.grid(row=0, column=1, padx=10, pady=5)

        # 오늘 이후 예정 공모주
        btn_upcoming_all = tk.Button(
            btn_frame,
            text="오늘 이후 예정 공모주 보기",
            width=25,
            command=self.show_upcoming_all,
        )
        btn_upcoming_all.grid(row=1, column=1, padx=10, pady=5)

        # 증권사별 보기
        btn_broker = tk.Button(
            btn_frame,
            text="증권사별 보기",
            width=20,
            command=self.open_broker_popup,
        )
        btn_broker.grid(row=0, column=2, padx=10, pady=5)

        # 🔥 프로그램 종료 버튼
        btn_exit = tk.Button(
            btn_frame,
            text="프로그램 종료",
            width=20,
            fg="white",
            bg="#D9534F",
            command=self.exit_program,
        )
        btn_exit.grid(row=1, column=2, padx=10, pady=5)

        # 로그 출력 Text
        self.text = tk.Text(self.root, font=("맑은 고딕", 11))
        self.text.pack(fill="both", expand=True, padx=10, pady=10)

        # 스크롤바
        scroll = tk.Scrollbar(self.text)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.text.config(yscrollcommand=scroll.set)
        scroll.config(command=self.text.yview)

    # ----------------------- 로그 유틸 -----------------------

    def log(self, msg: str):
        self.text.insert(tk.END, msg + "\n")
        self.text.see(tk.END)

    # ----------------------- 크롤링 스레드/컨트롤 -----------------------

    def run_collect_in_thread(self):
        """데이터 수집 버튼 클릭 → 별도 스레드에서 크롤링 실행"""
        # 중지 플래그 초기화
        self.stop_flag = False

        # 버튼 상태/라벨/스피너 설정
        self.btn_collect.config(state="disabled")
        self.loading_label.config(text="⏳ 데이터 수집 중… 잠시만 기다려주세요.")
        self.spinner_running = True
        self.animate_spinner()

        th = threading.Thread(target=self._collect_wrapper)
        th.daemon = True
        th.start()

    def _collect_wrapper(self):
        try:
            self.collect_data()
            if not self.stop_flag:
                self.loading_label.config(text="✅ 데이터 수집 완료!")
            else:
                self.loading_label.config(text="⛔ 크롤링이 중간에 중지되었습니다.")
        except Exception as e:
            self.loading_label.config(text=f"❌ 오류 발생: {e}")
        finally:
            self.spinner_running = False
            self.btn_collect.config(state="normal")

    def animate_spinner(self):
        if not self.spinner_running:
            return

        current = self.loading_label.cget("text")
        if "⏳" in current:
            new = current.replace("⏳", "🔄")
        else:
            new = current.replace("🔄", "⏳")
        self.loading_label.config(text=new)

        self.root.after(400, self.animate_spinner)

    def stop_crawling(self):
        """중지 버튼 → stop_flag만 True로 바꿔서 루프 종료 요청"""
        self.stop_flag = True
        self.loading_label.config(text="⛔ 크롤링 중지 요청됨…")

    # ----------------------- 프로그램 종료 -----------------------

    def exit_program(self):
        """프로그램 완전 종료"""
        self.root.destroy()

    # ----------------------- 기능 1: 데이터 수집 -----------------------

    def collect_data(self):
        self.log("=== 38커뮤니케이션 크롤링 시작 ===")
        try:
            total = crawl_38_all(
                log_func=self.log,
                stop_checker=lambda: self.stop_flag,
            )
            if not self.stop_flag:
                self.log(f"✅ 전체 {total}건 저장 완료")
        except Exception as e:
            self.log(f"❌ 오류 발생: {e}")

    # ----------------------- 기능 2: 엑셀(xlsx) 내보내기 -----------------------

    def export_to_excel(self):
        """SQLite 전체 데이터를 .xlsx(스타일 가능)로 저장"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        conn = get_connection()
        cur = conn.cursor()

        cur.execute(
            """
            SELECT stock_name, status, lead_manager, brokers, offer_price,
                sub_start, sub_end, listing_date, demand_start, demand_end,
                refund_date, source
            FROM ipo_schedules
            ORDER BY id
            """
        )
        rows = cur.fetchall()
        conn.close()

        if not rows:
            self.log("내보낼 데이터가 없습니다.")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "IPO Data"

        headers = [
            "종목명",
            "상태",
            "대표주관사",
            "증권사전체",
            "공모가",
            "청약시작일",
            "청약종료일",
            "상장일",
            "수요예측시작",
            "수요예측종료",
            "환불일",
            "출처",
        ]

        ws.append(headers)

        # 헤더 스타일
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        # 데이터 입력
        for row in rows:
            clean_source = row[11].split("_", 1)[-1] if row[11] else ""

            ws.append(
                [
                    row[0],  # 종목명
                    row[1],  # 상태
                    row[2],  # 대표주관사
                    row[3],  # 증권사전체
                    row[4],  # 공모가
                    row[5],  # 청약시작
                    row[6],  # 청약종료
                    row[7],  # 상장일
                    row[8],  # 수요예측시작
                    row[9],  # 수요예측종료
                    row[10],  # 환불일
                    clean_source,  # 🔥 숫자 prefix 제거된 출처
                ]
            )

        # 열 너비 조정
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 13

        for col in ["F", "G", "H", "I", "J", "K", "L"]:
            ws.column_dimensions[col].width = 15

        # 공모가(E열) 금액 서식
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            cell = row[0]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "₩#,##0"

        # 필터
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        # 🔥 바탕화면 경로 생성
        desktop = r"C:\Users\rhkdd\OneDrive\Desktop"
        self.log(f"[DEBUG] 실제 바탕화면 경로: {desktop}")
        path = os.path.join(desktop, "크롤링데이터.xlsx")
        try:
            wb.save(path)
            self.log(f"✅ 저장 성공: {path}")
        except Exception as e:
            self.log(f"❌ 저장 실패: {e}")
        # # Windows 실제 바탕화면 폴더 반환 (OneDrive도 자동 처리)
        # CSIDL_DESKTOP = 0x0000
        # SHGFP_TYPE_CURRENT = 0
        # buf = ctypes.create_unicode_buffer(260)
        # ctypes.windll.shell32.SHGetFolderPathW(None, CSIDL_DESKTOP, None, SHGFP_TYPE_CURRENT, buf)
        # desktop = buf.value

        # path = os.path.join(desktop, "크롤링데이터.xlsx")

        self.log(f"✅ 엑셀(xlsx) 파일 저장 완료: {path}")

    # ----------------------- 기능 3: 전체 예정 공모주 -----------------------

    def show_upcoming_all(self):
        """오늘 기준 이후의 모든 예정 공모주 출력"""
        conn = get_connection()
        cur = conn.cursor()

        today = datetime.now().strftime("%Y-%m-%d")

        query = """
        SELECT stock_name, status, sub_start, sub_end,
               demand_start, demand_end, listing_date, source
        FROM ipo_schedules
        WHERE
            (sub_start IS NOT NULL AND sub_start >= ?) OR
            (demand_start IS NOT NULL AND demand_start >= ?) OR
            (listing_date IS NOT NULL AND listing_date >= ?)
        ORDER BY
            CASE
                WHEN sub_start IS NOT NULL THEN sub_start
                WHEN demand_start IS NOT NULL THEN demand_start
                WHEN listing_date IS NOT NULL THEN listing_date
            END
        """

        cur.execute(query, (today, today, today))
        rows = cur.fetchall()
        conn.close()

        self.log("")
        self.log(f"=== 오늘({today}) 기준 예정 공모주 ===")

        if not rows:
            self.log("예정 공모주가 없습니다.")
            return

        for stock, status, ss, se, ds, de, ld, source in rows:
            if ss:
                date_str = f"{ss} ~ {se}"
            elif ds:
                date_str = f"{ds} ~ {de}"
            elif ld:
                date_str = ld
            else:
                date_str = "-"

            self.log(f"- {stock} [{status}] {date_str} (출처: {source})")

    # ----------------------- 기능 4: 증권사별 보기 -----------------------

    def _get_all_brokers(self):
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT brokers FROM ipo_schedules")
        rows = cur.fetchall()
        conn.close()

        broker_set = set()
        for (b,) in rows:
            if not b:
                continue
            parts = b.split(",")
            for p in parts:
                p = p.strip()
                if not p:
                    continue
                if any(key in p for key in ["증권", "투자", "스팩"]):
                    broker_set.add(p)

        return sorted(broker_set)

    def _show_broker_result(self, broker_name: str):
        rows = get_upcoming_by_broker(broker_name)

        today = datetime.now().strftime("%Y-%m-%d")
        self.log("")
        self.log(f"=== {broker_name} 기준 예정 공모주 (오늘 {today} 이후) ===")

        if not rows:
            self.log("해당 증권사가 주관하는 예정 공모주가 없습니다.")
            return

        for stock, status, ss, se, ds, de, ld, source in rows:
            if ss:
                date_str = f"{ss} ~ {se}"
            elif ds:
                date_str = f"{ds} ~ {de}"
            elif ld:
                date_str = ld
            else:
                date_str = "-"

            self.log(f"- {stock} [{status}] {date_str} (출처: {source})")

    def open_broker_popup(self):
        popup = tk.Toplevel(self.root)
        popup.title("증권사 선택")
        popup.geometry("350x400")

        tk.Label(popup, text="증권사를 선택하세요:", font=("맑은 고딕", 12)).pack(
            pady=10
        )

        listbox = tk.Listbox(popup, width=30, height=15)
        listbox.pack(pady=5)

        brokers = self._get_all_brokers()
        for b in brokers:
            listbox.insert(tk.END, b)

        def select_broker():
            sel = listbox.curselection()
            if not sel:
                return
            broker = listbox.get(sel[0])
            popup.destroy()
            self._show_broker_result(broker)

        tk.Button(popup, text="선택", command=select_broker).pack(pady=10)
